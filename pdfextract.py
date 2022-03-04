# This Python file uses the following encoding: utf-8
import pdfplumber
import pandas as pd
import openpyxl
import openpyxl.styles
from openpyxl.styles import Font, Border, Alignment
import time
from datetime import date
import json
import platform


if platform == "linux":
    json_file_name = "config.json"
else:
    json_file_name = "config.json"

def multiple_pdfs_to_txt(files):
    """Converts all information in the files to string and transfer it to excel file"""
    json_file = open(json_file_name, encoding="utf8")
    json_data = json.load(json_file)
    excel_file = pd.DataFrame()
    excel_file_name = files[0][:-4] + ".xlsx"
    excel_file.to_excel(excel_file_name)

    if platform == "linux":
        book = openpyxl.load_workbook("" + excel_file_name)
    else:
        book = openpyxl.load_workbook(json_data["path"] + excel_file_name)
    sheet = book.active

    sheet.sheet_view.rightToLeft = True

    have_found_file_type = False

    people_row_count = json_data["excel_row_information_start"]
    company_row_count = json_data["excel_row_information_start"]
    passport_row_count = json_data["excel_row_information_start"]

    update_text_file = files[0][:-4]+ ".txt"
    file_type = 0
    added_row = 0
    for file in files:
        pdf_file_name = file[:-24] + ".pdf"
        last_line = int(max(people_row_count, company_row_count, passport_row_count))
        sheet.cell(row=last_line, column=1).value = "נלקח ממסמך: " + str(pdf_file_name)

        people_row_count= int(last_line + 1)
        company_row_count= int(last_line + 1)
        passport_row_count = int(last_line + 1)

        with pdfplumber.open(pdf_file_name) as pdf:
            for page in pdf.pages:
                val = str(page)[1:]
                val = val[:-1]
                page_num = val[5:]
                temp_page_amount = str(pdf.pages)[-11:]
                page_amount = ""
                for char in temp_page_amount:
                    if char.isnumeric():
                        page_amount += char
                page_amount = " out of  " + page_amount

                val = "Page " + page_num + page_amount
                write_data_in_file_multiple_pdfs(val, update_text_file, pdf_file_name)
                for line in page.extract_text().split("\n"):
                    if have_found_file_type:
                        added_row = line_information_extractor(
                            line,
                            file_type,
                            sheet,
                            people_row_count,
                            company_row_count,
                            passport_row_count,
                            page_num,
                        )
                        if added_row == 1:
                            people_row_count += 1
                        elif added_row == 2:
                            company_row_count += 1
                        elif added_row == 3:
                            passport_row_count += 1
                    else:
                        file_type = find_file_type(line, sheet)
                        if file_type != 0:
                            have_found_file_type = True

        # Adding titles
    write_excel_titles(sheet)

    # Saving the excel

    new_excel_file_name = excel_file_name[:-5] + " result.xlsx"
    book.title = new_excel_file_name
    book.save(new_excel_file_name)

    # Adding the information to the information file
    write_data_in_file("Finished extracting " + files[0][:-4], files[0][:-4] + ".txt")


def pdf_to_txt(file):
    """Converting every line in the pdf into a line into an excel it created with the name of the pdf + "result" """

    json_file = open(json_file_name, encoding="utf8")
    json_data = json.load(json_file)
    excel_file = pd.DataFrame()
    excel_file_name = file[:-4] + ".xlsx"
    excel_file.to_excel(excel_file_name)

    if platform == "linux":
        book = openpyxl.load_workbook("" + excel_file_name)
    else:
        book = openpyxl.load_workbook(json_data["path"] + excel_file_name)
    sheet = book.active

    sheet.sheet_view.rightToLeft = True

    have_found_file_type = False

    people_row_count = json_data["excel_row_information_start"]
    company_row_count = json_data["excel_row_information_start"]
    passport_row_count = json_data["excel_row_information_start"]

    file_type = 0
    added_row = 0

    pdf_file_name = file[:-24] + ".pdf"
    with pdfplumber.open(pdf_file_name) as pdf:
        for page in pdf.pages:
            val = str(page)[1:]
            val = val[:-1]
            page_num = val[5:]
            temp_page_amount = str(pdf.pages)[-11:]
            page_amount = ""
            for char in temp_page_amount:
                if char.isnumeric():
                    page_amount += char
            page_amount = " out of  " + page_amount

            val = "Page " + page_num + page_amount
            write_data_in_file(val, file[:-4] + ".txt")
            for line in page.extract_text().split("\n"):
                if have_found_file_type:
                    added_row = line_information_extractor(
                        line,
                        file_type,
                        sheet,
                        people_row_count,
                        company_row_count,
                        passport_row_count,
                        page_num,
                    )
                    if added_row == 1:
                        people_row_count += 1
                    elif added_row == 2:
                        company_row_count += 1
                    elif added_row == 3:
                        passport_row_count += 1
                else:
                    file_type = find_file_type(line, sheet)
                    if file_type != 0:
                        have_found_file_type = True

        # Adding titles
    write_excel_titles(sheet)

    # Saving the excel

    new_excel_file_name = excel_file_name[:-5] + " result.xlsx"
    book.title = new_excel_file_name
    book.save(new_excel_file_name)

    # Adding the information to the information file
    write_data_in_file("Finished extracting " + file[:-4], file[:-4] + ".txt")


def line_information_extractor(
    info,
    type_of_file,
    sheet,
    people_row_count,
    company_row_count,
    passport_row_count,
    page,
):
    """getting a line and checking if a certain information is in it then writing it in the excel"""
    json_file = open(json_file_name, encoding="utf8")
    json_data = json.load(json_file)

    row_added = 0

    if isinstance(info, str):
        # Checking if there`s ID in the line
        if json_data["hebrew_ID"] in info:
            info = info + " "
            info = " ".join(info.split())
            info = " " + info
            print(info)

            if type_of_file == 1:
                id_value = get_ID_from_sentence(info)
                id_value = id_value.replace(" ", "")

                sheet.cell(row=people_row_count, column=1).value = id_value
                sheet.cell(row=people_row_count, column=1).font = Font(
                    size=11, bold=False
                )

                # Findind the name and putting it in the excel by certain distance from the ID and the reason
                name_value = get_ID_name_from_sentence(info)[::-1]
                name_value = name_value.replace("  ", " ")

                sheet.cell(row=people_row_count, column=2).value = name_value
                sheet.cell(row=people_row_count, column=2).font = Font(
                    size=11, bold=False
                )

                # Printing for debugging
                print(id_value)
                print(name_value)

            if type_of_file == 2:
                # Find the ID
                id_value = id_value = get_ID_from_sentence(info)
                id_value = id_value.replace(" ", "")

                sheet.cell(
                    row=people_row_count,
                    column=1,
                ).value = id_value
                sheet.cell(row=people_row_count, column=1).font = Font(
                    size=11, bold=False
                )

                # Finding the name
                name_value = get_ID_name_from_sentence(info)[::-1]
                name_value = name_value.replace("  ", " ")

                sheet.cell(row=people_row_count, column=2).value = name_value
                sheet.cell(row=people_row_count, column=2).font = Font(
                    size=11, bold=False
                )

                # Printing for debugging
                print(id_value)
                print(name_value)

            if type_of_file == 3:
                # Find the ID
                id_value = id_value = get_ID_from_sentence(info)
                id_value = id_value.replace(" ", "")

                sheet.cell(
                    row=people_row_count,
                    column=1,
                ).value = id_value
                sheet.cell(row=people_row_count, column=1).font = Font(
                    size=11, bold=False
                )

                # Finding the name
                name_value = get_ID_name_from_sentence(info)[::-1]
                name_value = name_value.replace("  ", " ")

                sheet.cell(row=people_row_count, column=2).value = name_value
                sheet.cell(row=people_row_count, column=2).font = Font(
                    size=11, bold=False
                )
            # Adding 1 to the index of where the program will write
            sheet.cell(row=people_row_count, column=3).value = page
            sheet.cell(row=people_row_count, column=3).font = Font(size=11, bold=False)
            row_added = 1
        # Checking if there`s company and not mortgage in the line
        elif (
            json_data["hebrew_Company"] in info
            and json_data["hebrew_Mortgage"] not in info
        ):
            info += " "
            info = " " + info
            info = " ".join(info.split())
            print(info)

            if type_of_file == 1:
                # Finding the company ID and putting it in the excel
                company_id_value = get_ID_from_sentence(info)
                if company_id_value == None:
                    return 0
                company_id_value = company_id_value.replace(" ", "")

                sheet.cell(row=company_row_count, column=5).value = company_id_value
                sheet.cell(row=company_row_count, column=5).font = Font(
                    size=11, bold=False
                )

                # Finding the name
                company_name_value = get_company_name_from_sentence(info)

                sheet.cell(row=company_row_count, column=6).value = company_name_value
                sheet.cell(row=company_row_count, column=6).font = Font(
                    size=11, bold=False
                )

                # Printing for debugging
                print(company_name_value)
                print(company_id_value)

            if type_of_file == 2:
                # Finding the company ID and putting it in the excel
                company_id_value = get_ID_from_sentence(info)
                if company_id_value == None:
                    return 0
                company_id_value = company_id_value.replace(" ", "")

                sheet.cell(row=company_row_count, column=5).value = company_id_value
                sheet.cell(row=company_row_count, column=5).font = Font(
                    size=11, bold=False
                )

                # Finding the name
                company_name_value = get_company_name_from_sentence(info)

                sheet.cell(row=company_row_count, column=6).value = company_name_value
                sheet.cell(row=company_row_count, column=6).font = Font(
                    size=11, bold=False
                )

                # Printing for debugging
                print(company_name_value)
                print(company_id_value)
            if type_of_file == 3:
                # Finding the company ID and putting it in the excel
                company_id_value = get_ID_from_sentence(info)
                if company_id_value == None:
                    return 0
                company_id_value = company_id_value.replace(" ", "")

                sheet.cell(row=company_row_count, column=5).value = company_id_value
                sheet.cell(row=company_row_count, column=5).font = Font(
                    size=11, bold=False
                )

                # Finding the name
                company_name_value = get_company_name_from_sentence(info)

                sheet.cell(row=company_row_count, column=6).value = company_name_value
                sheet.cell(row=company_row_count, column=6).font = Font(
                    size=11, bold=False
                )

                # Printing for debugging
                print(company_name_value)
                print(company_id_value)

            sheet.cell(row=company_row_count, column=7).value = page
            sheet.cell(row=company_row_count, column=7).font = Font(size=11, bold=False)

            # Adding 1 to the index of where the program will write
            row_added = 2

        # Checking if there`s passport in the line
        elif json_data["hebrew_passport"] in info:
            info += " "
            info = " " + info
            info = " ".join(info.split())
            print(info)

            if type_of_file == 1:
                # Find the passport and putting it in the excel (more complicated check, ID comes in multiple lengths)
                passport_value = get_passport_from_sentence(info)
                if passport_value == None:
                    return 0
                passport_value = passport_value.replace(" ", "")

                sheet.cell(row=passport_row_count, column=8).value = passport_value
                sheet.cell(row=passport_row_count, column=8).font = Font(
                    size=11, bold=False
                )

                passport_name_value = get_passport_name_from_sentence(info)[::-1]
                sheet.cell(row=passport_row_count, column=9).value = passport_name_value
                sheet.cell(row=passport_row_count, column=9).font = Font(
                    size=11, bold=False
                )

                print(passport_value)
                print(passport_name_value)

                # Printing for debugging
                print(passport_value)
                print(passport_name_value)
            if type_of_file == 2:
                passport_value = get_passport_from_sentence(info)
                if passport_value == None:
                    return 0
                sheet.cell(row=passport_row_count, column=8).value = passport_value
                sheet.cell(row=passport_row_count, column=8).font = Font(
                    size=11, bold=False
                )

                passport_name_value = get_passport_name_from_sentence(info)
                sheet.cell(row=passport_row_count, column=9).value = passport_name_value
                sheet.cell(row=passport_row_count, column=9).font = Font(
                    size=11, bold=False
                )

                print(passport_value)
                print(passport_name_value)
                
            if type_of_file == 3:
                passport_value = get_passport_from_sentence(info)
                if passport_value == None:
                    return 0
                sheet.cell(row=passport_row_count, column=8).value = passport_value
                sheet.cell(row=passport_row_count, column=8).font = Font(
                    size=11, bold=False
                )

                passport_name_value = get_passport_name_from_sentence(info)
                sheet.cell(row=passport_row_count, column=9).value = passport_name_value
                sheet.cell(row=passport_row_count, column=9).font = Font(
                    size=11, bold=False
                )

                print(passport_value)
                print(passport_name_value)
            # Adding 1 to the index of where the program will write
            sheet.cell(row=passport_row_count, column=10).value = page
            sheet.cell(row=passport_row_count, column=10).font = Font(
                size=11, bold=False
            )
            row_added = 3

    return row_added


def write_data_in_file_multiple_pdfs(value, filename, pdf_file_name):
    """writes in the information file the file name, and time of executing"""
    print(filename)
    information_file = open(filename, "a")
    information_file.write(
        value
        + " "
        + str(date.today().strftime("%d/%m/%Y"))
        + " "
        + str(time.strftime("%H:%M:%S", time.localtime()))
        + "   - "
        + pdf_file_name
        + "\n"
    )
    information_file.close()
    write_data_in_information_file(value)

def write_data_in_file(value, filename):
    """writes in the information file the file name, and time of executing"""
    print(filename)
    information_file = open(filename, "a")
    information_file.write(
        value
        + " "
        + str(date.today().strftime("%d/%m/%Y"))
        + " "
        + str(time.strftime("%H:%M:%S", time.localtime()))
        + "\n"
    )
    information_file.close()


def write_data_in_information_file(value):
    filename = "Information.txt"
    information_file = open(filename, "a")
    information_file.write(
        value
        + " "
        + str(date.today().strftime("%d/%m/%Y"))
        + " "
        + str(time.strftime("%H:%M:%S", time.localtime()))
        + "\n"
    )
    information_file.close()


def write_excel_titles(sheet):
    """gets a sheet and writes titles in it"""
    sheet.cell(row=1, column=1).value = "ת.ז"
    sheet.cell(row=1, column=1).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=2).value = "שם"
    sheet.cell(row=1, column=2).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=3).value = "מספר עמוד"
    sheet.cell(row=1, column=3).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=5).value = "מספר"
    sheet.cell(row=1, column=5).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=6).value = "שם חברה"
    sheet.cell(row=1, column=6).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=7).value = "מספר עמוד"
    sheet.cell(row=1, column=7).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=8).value = "מספר דרכון"
    sheet.cell(row=1, column=8).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=9).value = "שם"
    sheet.cell(row=1, column=9).font = Font(size=11, bold=True)
    sheet.cell(row=1, column=10).value = "מספר עמוד"
    sheet.cell(row=1, column=10).font = Font(size=11, bold=True)


def clear_excel_cell(cell):
    """receive excel cell and clears it"""
    cell.value = None
    cell.font = Font(size=11, bold=False)
    cell.border = Border()


def write_file_type_in_excel(file_type, sheet):
    """gets sheet and file-type and writes it in the sheet"""
    sheet.cell(row=1, column=11).value = "סוג קובץ:"
    sheet.cell(row=1, column=11).font = Font(size=11, bold=True)
    sheet.cell(row=2, column=11).value = file_type
    sheet.cell(row=2, column=11).font = Font(size=11, bold=False)


def find_file_type(info, sheet):
    """Returning the type of the file presented by numbers"""
    if "םיפתושמ םיתב" in info:
        print(info)
        write_file_type_in_excel("בתים משותפים", sheet)
        return 1
    elif "תויוכזה סקנפמ" in info:
        print(info)
        write_file_type_in_excel("פנקס זכויות", sheet)
        return 2
    elif "תורטשה סקנפמ" in info:
        print(info)
        write_file_type_in_excel("פנקס השטרות", sheet)
        return 3
    else:
        return 0


def get_ID_from_sentence(sentence):
    words = sentence.split()
    for word in words:
        if len(word) > 5:
            if word.isnumeric() or ("/" not in word and "-" in word):
                return word


def get_ID_name_from_sentence(info):
    """Returning the name"""
    json_file = open(json_file_name, encoding="utf8")
    json_data = json.load(json_file)

    info = info.replace(json_data["hebrew_ID"], "")

    info = info[::-1]

    for reason in json_data["possible_name_reasons"]:
        if reason in info:
            info = info.replace(reason, "")

    for word in info:
        if word.isdigit() or "/" in word:
            info = info.replace(word, "")

    for count in range(0, len(info)):
        if info[count] == ")":
            info = info[:count] + "(" + info[count + 1 :]
        elif info[count] == "(":
            info = info[:count] + ")" + info[count + 1 :]

    info = info.replace("  ", "")
    for a in range(0, 2):
        if info[len(info) - 1] == " " or info[len(info) - 1] == "-":
            info = info[:-1]
        elif info[0] == " " or info[0] == "-":
            info = info[1:]

    info = info[::-1]

    return info


def get_company_name_from_sentence(info):
    json_file = open(json_file_name, encoding="utf8")
    json_data = json.load(json_file)

    info = info.replace(json_data["hebrew_Company"], "")
    info = info[::-1]

    for reason in json_data["possible_company_name_reasons"]:
        if reason in info:
            info = info.replace(reason, "")

    for word in info:
        if word.isdigit() or "/" in word:
            info = info.replace(word, "")

    for count in range(0, len(info)):
        if info[count] == ")":
            info = info[:count] + "(" + info[count + 1 :]
        elif info[count] == "(":
            info = info[:count] + ")" + info[count + 1 :]

    info = info.replace("  ", "")
    for a in range(0, 2):
        if info[len(info) - 1] == " " or info[len(info) - 1] == "-":
            info = info[:-1]
        elif info[0] == " " or info[0] == "-":
            info = info[1:]

    return info


def get_passport_from_sentence(sentence):
    words = sentence.split()
    for word in words:
        isGood = True
        if len(word) > 5:
            if word.isdigit() or word.isupper():
                return word
            else:
                for char in word:
                    if not char.isdigit() and not char.isupper() and not char.islower():
                        isGood = False
                if isGood:
                    return word


def get_passport_name_from_sentence(info):
    """Returning the name"""
    json_file = open(json_file_name, encoding="utf8")
    json_data = json.load(json_file)

    info = info.replace(json_data["hebrew_passport"], "")

    info = info[::-1]

    for reason in json_data["possible_name_reasons"]:
        if reason in info:
            info = info.replace(reason, "")

    for word in info:
        if word.isdigit() or "/" in word or word.isupper() or word.islower():
            info = info.replace(word, "")

    for count in range(0, len(info)):
        if info[count] == ")":
            info = info[:count] + "(" + info[count + 1 :]
        elif info[count] == "(":
            info = info[:count] + ")" + info[count + 1 :]

    return info


# pdf_to_txt('352.pdf')
